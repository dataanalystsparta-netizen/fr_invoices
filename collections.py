# ==========================================================
# FASTRANKING COLLECTIONS DASHBOARD
# PART 1A - IMPORTS, CONFIGURATION & DATA LOADING
# ==========================================================

import streamlit as st
import pandas as pd
import numpy as np
import sqlite3
from datetime import datetime


# ----------------------------------------------------------
# PAGE CONFIGURATION
# ----------------------------------------------------------

st.set_page_config(
    page_title="FastRanking Collections Dashboard",
    page_icon="📞",
    layout="wide"
)

st.title("📞 FastRanking Collections Dashboard")


# ----------------------------------------------------------
# SOURCE FILES
# ----------------------------------------------------------

INVOICE_FILE = "Invoice_zoho.xlsx"
PAYMENT_FILE = "Customer_Payment_zoho.xlsx"
CONTACTS_FILE = "Contacts_zoho.xlsx"


# ----------------------------------------------------------
# HELPER FUNCTIONS
# ----------------------------------------------------------

def clean_text_series(series):
    """
    Convert values to clean text while preserving blanks.
    Prevents values such as NaN from becoming visible strings.
    """

    return (
        series
        .fillna("")
        .astype(str)
        .str.strip()
    )


def first_non_blank(series):
    """
    Return the first non-blank value from a grouped column.
    """

    values = series.dropna().astype(str).str.strip()
    values = values[values != ""]

    if values.empty:
        return ""

    return values.iloc[0]



############################################################
## DB_FUNCS
############################################################
DB_FILE = "collections_updates.db"


def load_collection_updates():

    conn = sqlite3.connect(DB_FILE)

    query = """
    SELECT *
    FROM collection_updates
    """

    try:
        df = pd.read_sql(
            query,
            conn
        )

    except:

        df = pd.DataFrame(
            columns=[
                "Customer Name",
                "Disposition",
                "PTP Date",
                "Remarks"
            ]
        )

    conn.close()

    return df



def save_collection_updates(df):

    conn = sqlite3.connect(DB_FILE)

    df["Updated At"] = datetime.now()

    df.to_sql(
        "collection_updates",
        conn,
        if_exists="replace",
        index=False
    )

    conn.close()

# ==========================================================
# LOAD AND PREPARE DATA
# ==========================================================

@st.cache_data(show_spinner=False)
def load_data():

    # ------------------------------------------------------
    # READ EXCEL FILES
    # ------------------------------------------------------

    invoices = pd.read_excel(INVOICE_FILE)
    payments = pd.read_excel(PAYMENT_FILE)
    contacts = pd.read_excel(CONTACTS_FILE)

    # ------------------------------------------------------
    # CLEAN COLUMN NAMES
    # ------------------------------------------------------

    invoices.columns = invoices.columns.astype(str).str.strip()
    payments.columns = payments.columns.astype(str).str.strip()
    contacts.columns = contacts.columns.astype(str).str.strip()

    # ------------------------------------------------------
    # REQUIRED COLUMN CHECKS
    # ------------------------------------------------------

    required_invoice_columns = [
        "Invoice Number",
        "Invoice Date",
        "Due Date",
        "Invoice Status",
        "Customer Name",
        "Total",
        "Balance"
    ]

    required_payment_columns = [
        "Invoice Number",
        "Amount Applied to Invoice",
        "Date"
    ]

    required_contact_columns = [
        "Display Name"
    ]

    missing_invoice_columns = [
        col
        for col in required_invoice_columns
        if col not in invoices.columns
    ]

    missing_payment_columns = [
        col
        for col in required_payment_columns
        if col not in payments.columns
    ]

    missing_contact_columns = [
        col
        for col in required_contact_columns
        if col not in contacts.columns
    ]

    if missing_invoice_columns:
        raise ValueError(
            "Missing invoice columns: "
            + ", ".join(missing_invoice_columns)
        )

    if missing_payment_columns:
        raise ValueError(
            "Missing payment columns: "
            + ", ".join(missing_payment_columns)
        )

    if missing_contact_columns:
        raise ValueError(
            "Missing contact columns: "
            + ", ".join(missing_contact_columns)
        )

    # ------------------------------------------------------
    # CLEAN MATCHING COLUMNS
    # ------------------------------------------------------

    invoices["Invoice Number"] = clean_text_series(
        invoices["Invoice Number"]
    )

    invoices["Customer Name"] = clean_text_series(
        invoices["Customer Name"]
    )

    invoices["Invoice Status"] = clean_text_series(
        invoices["Invoice Status"]
    )

    payments["Invoice Number"] = clean_text_series(
        payments["Invoice Number"]
    )

    contacts["Display Name"] = clean_text_series(
        contacts["Display Name"]
    )

    # ------------------------------------------------------
    # CONVERT INVOICE DATES
    # ------------------------------------------------------

    invoice_date_columns = [
        "Invoice Date",
        "Due Date",
        "Last Payment Date",
        "Expected Payment Date"
    ]

    for col in invoice_date_columns:

        if col in invoices.columns:

            invoices[col] = pd.to_datetime(
                invoices[col],
                dayfirst=True,
                errors="coerce"
            )

    # ------------------------------------------------------
    # CONVERT PAYMENT DATES
    # ------------------------------------------------------

    payments["Date"] = pd.to_datetime(
        payments["Date"],
        dayfirst=True,
        errors="coerce"
    )

    if "Invoice Date" in payments.columns:

        payments["Invoice Date"] = pd.to_datetime(
            payments["Invoice Date"],
            dayfirst=True,
            errors="coerce"
        )

    # ------------------------------------------------------
    # CONVERT CONTACT DATES
    # ------------------------------------------------------

    for col in [
        "Created Time",
        "Last Modified Time"
    ]:

        if col in contacts.columns:

            contacts[col] = pd.to_datetime(
                contacts[col],
                errors="coerce"
            )

    # ------------------------------------------------------
    # CONVERT INVOICE NUMERIC COLUMNS
    # ------------------------------------------------------

    for col in [
        "Total",
        "Balance",
        "SubTotal"
    ]:

        if col in invoices.columns:

            invoices[col] = (
                pd.to_numeric(
                    invoices[col],
                    errors="coerce"
                )
                .fillna(0)
            )

    # ------------------------------------------------------
    # CONVERT PAYMENT NUMERIC COLUMNS
    # ------------------------------------------------------

    for col in [
        "Amount",
        "Amount Applied to Invoice"
    ]:

        if col in payments.columns:

            payments[col] = (
                pd.to_numeric(
                    payments[col],
                    errors="coerce"
                )
                .fillna(0)
            )

    # ------------------------------------------------------
    # REMOVE DRAFT AND VOID INVOICES
    # ------------------------------------------------------

    invoices = invoices[
        ~invoices["Invoice Status"]
        .str.casefold()
        .isin(["draft", "void"])
    ].copy()

    # ------------------------------------------------------
    # REMOVE BLANK INVOICE NUMBERS
    # ------------------------------------------------------

    invoices = invoices[
        invoices["Invoice Number"] != ""
    ].copy()

    # ------------------------------------------------------
    # REMOVE DUPLICATE INVOICE NUMBERS
    # ------------------------------------------------------

    invoices = (
        invoices
        .sort_values(
            by=["Invoice Date", "Invoice Number"],
            na_position="last"
        )
        .drop_duplicates(
            subset="Invoice Number",
            keep="first"
        )
        .reset_index(drop=True)
    )

    # ------------------------------------------------------
    # BUILD PAYMENT SUMMARY
    # ------------------------------------------------------

    payment_summary = (
        payments[
            payments["Invoice Number"] != ""
        ]
        .groupby(
            "Invoice Number",
            as_index=False
        )
        .agg(
            Paid=(
                "Amount Applied to Invoice",
                "sum"
            ),
            Last_Payment=(
                "Date",
                "max"
            )
        )
    )

    # ------------------------------------------------------
    # MERGE PAYMENTS INTO INVOICES
    # ------------------------------------------------------

    invoices = invoices.merge(
        payment_summary,
        on="Invoice Number",
        how="left",
        validate="one_to_one"
    )

    invoices["Paid"] = (
        pd.to_numeric(
            invoices["Paid"],
            errors="coerce"
        )
        .fillna(0)
    )

    # Prevent payment values from exceeding invoice totals
    invoices["Paid"] = invoices[
        ["Paid", "Total"]
    ].min(axis=1)

    # ------------------------------------------------------
    # KEEP OUTSTANDING INVOICES ONLY
    # ------------------------------------------------------

    invoices = invoices[
        invoices["Balance"] > 0
    ].copy()

    # ------------------------------------------------------
    # CALCULATE STATUS
    # ------------------------------------------------------

    today = pd.Timestamp.today().normalize()

    invoices["Status"] = "Current"

    invoices.loc[
        invoices["Due Date"] < today,
        "Status"
    ] = "Overdue"

    invoices.loc[
        (invoices["Paid"] > 0)
        & (invoices["Balance"] > 0),
        "Status"
    ] = "Partially Paid"

    # ------------------------------------------------------
    # CALCULATE DAYS OVERDUE
    # ------------------------------------------------------

    invoices["Days Overdue"] = np.where(
        invoices["Due Date"] < today,
        (today - invoices["Due Date"]).dt.days,
        0
    )

    invoices["Days Overdue"] = (
        pd.to_numeric(
            invoices["Days Overdue"],
            errors="coerce"
        )
        .fillna(0)
        .clip(lower=0)
        .astype(int)
    )

    # ------------------------------------------------------
    # CALCULATE CURRENT AND OVERDUE BALANCES
    # ------------------------------------------------------

    invoices["Current Due"] = np.where(
        invoices["Due Date"] >= today,
        invoices["Balance"],
        0
    )

    invoices["Overdue Due"] = np.where(
        invoices["Due Date"] < today,
        invoices["Balance"],
        0
    )

    # ------------------------------------------------------
    # PREPARE CONTACT LOOKUP
    # ------------------------------------------------------

    contact_columns = [
        "Display Name",
        "Phone",
        "MobilePhone",
        "EmailID"
    ]

    for col in contact_columns:

        if col not in contacts.columns:
            contacts[col] = ""

        contacts[col] = clean_text_series(
            contacts[col]
        )

    contacts_lookup = (
        contacts[contact_columns]
        .groupby(
            "Display Name",
            as_index=False
        )
        .agg(
            Phone=("Phone", first_non_blank),
            MobilePhone=("MobilePhone", first_non_blank),
            EmailID=("EmailID", first_non_blank)
        )
    )

    # ------------------------------------------------------
    # MERGE CONTACT DETAILS
    # ------------------------------------------------------

    invoices = invoices.merge(
        contacts_lookup,
        left_on="Customer Name",
        right_on="Display Name",
        how="left",
        validate="many_to_one"
    )

    invoices = invoices.drop(
        columns=["Display Name"],
        errors="ignore"
    )

    for col in [
        "Phone",
        "MobilePhone",
        "EmailID"
    ]:

        invoices[col] = clean_text_series(
            invoices[col]
        )

    # Use mobile as the primary calling number when available
    invoices["Phone Number"] = np.where(
        invoices["MobilePhone"] != "",
        invoices["MobilePhone"],
        invoices["Phone"]
    )

    return invoices


# ----------------------------------------------------------
# LOAD PREPARED OUTSTANDING INVOICES
# ----------------------------------------------------------

calling_df = load_data()
# ==========================================================
# FASTRANKING COLLECTIONS DASHBOARD
# PART 1B - CUSTOMER SUMMARY & KPI CALCULATIONS
# ==========================================================

# ----------------------------------------------------------
# SAFETY CHECK
# ----------------------------------------------------------

if calling_df.empty:

    st.warning(
        "No outstanding invoices were found in the current files."
    )

    st.stop()


# ----------------------------------------------------------
# REMOVE ANY DUPLICATE INVOICE NUMBERS
# ----------------------------------------------------------

calling_df = (
    calling_df
    .sort_values(
        by=["Invoice Date", "Invoice Number"],
        na_position="last"
    )
    .drop_duplicates(
        subset="Invoice Number",
        keep="first"
    )
    .reset_index(drop=True)
)


# ----------------------------------------------------------
# CUSTOMER-LEVEL COLLECTIONS SUMMARY
# ----------------------------------------------------------

customer_table = (
    calling_df
    .groupby(
        "Customer Name",
        as_index=False
    )
    .agg(
        Phone_Number=(
            "Phone Number",
            first_non_blank
        ),
        Phone=(
            "Phone",
            first_non_blank
        ),
        Mobile=(
            "MobilePhone",
            first_non_blank
        ),
        Email=(
            "EmailID",
            first_non_blank
        ),
        Invoice_Count=(
            "Invoice Number",
            "nunique"
        ),
        Total_Due=(
            "Balance",
            "sum"
        ),
        Current_Due=(
            "Current Due",
            "sum"
        ),
        Overdue_Due=(
            "Overdue Due",
            "sum"
        ),
        Total_Invoiced=(
            "Total",
            "sum"
        ),
        Total_Paid=(
            "Paid",
            "sum"
        ),
        Oldest_Due_Date=(
            "Due Date",
            "min"
        ),
        Latest_Invoice_Date=(
            "Invoice Date",
            "max"
        ),
        Maximum_Days_Overdue=(
            "Days Overdue",
            "max"
        ),
        Partially_Paid_Invoices=(
            "Status",
            lambda values: (
                values == "Partially Paid"
            ).sum()
        )
    )
)


# ----------------------------------------------------------
# CUSTOMER COLLECTION STATUS
# ----------------------------------------------------------

customer_table["Status"] = "Current"

customer_table.loc[
    customer_table["Overdue_Due"] > 0,
    "Status"
] = "Overdue"

customer_table.loc[
    customer_table["Partially_Paid_Invoices"] > 0,
    "Status"
] = "Partially Paid"


# ----------------------------------------------------------
# CUSTOMER PRIORITY
# ----------------------------------------------------------

customer_table["Priority"] = "🟢 Low"

customer_table.loc[
    customer_table["Maximum_Days_Overdue"] >= 30,
    "Priority"
] = "🟡 Medium"

customer_table.loc[
    customer_table["Maximum_Days_Overdue"] >= 90,
    "Priority"
] = "🔴 High"

# Partially paid customers need follow-up,
# unless they are already high priority.
customer_table.loc[
    (customer_table["Partially_Paid_Invoices"] > 0)
    & (customer_table["Maximum_Days_Overdue"] < 90),
    "Priority"
] = "🟦 Follow Up"


# ----------------------------------------------------------
# BLANK COLLECTION WORK COLUMNS
# ----------------------------------------------------------

customer_table["Disposition"] = ""
customer_table["PTP Date"] = ""
customer_table["Remarks"] = ""



################################################################
saved_updates = load_collection_updates()


if not saved_updates.empty:

    customer_table = customer_table.merge(
        saved_updates,
        on="Customer Name",
        how="left",
        suffixes=("", "_saved")
    )


    for col in [
        "Disposition",
        "PTP Date",
        "Remarks"
    ]:

        customer_table[col] = (
            customer_table[f"{col}_saved"]
            .fillna(customer_table[col])
        )


    customer_table = customer_table.drop(
        columns=[
            "Disposition_saved",
            "PTP Date_saved",
            "Remarks_saved"
        ],
        errors="ignore"
    )


# ----------------------------------------------------------
# SORT CUSTOMER QUEUE
# ----------------------------------------------------------

priority_order = {
    "🔴 High": 0,
    "🟦 Follow Up": 1,
    "🟡 Medium": 2,
    "🟢 Low": 3
}

customer_table["Priority Sort"] = (
    customer_table["Priority"]
    .map(priority_order)
    .fillna(99)
)

customer_table = (
    customer_table
    .sort_values(
        by=[
            "Priority Sort",
            "Maximum_Days_Overdue",
            "Total_Due",
            "Customer Name"
        ],
        ascending=[
            True,
            False,
            False,
            True
        ]
    )
    .drop(
        columns="Priority Sort"
    )
    .reset_index(drop=True)
)


# ==========================================================
# KPI CALCULATIONS
# ==========================================================

total_customers = (
    calling_df["Customer Name"]
    .nunique()
)

total_invoices = (
    calling_df["Invoice Number"]
    .nunique()
)

total_current_due = (
    calling_df["Current Due"]
    .sum()
)

total_overdue_due = (
    calling_df["Overdue Due"]
    .sum()
)

total_partial_due = (
    calling_df.loc[
        calling_df["Status"] == "Partially Paid",
        "Balance"
    ]
    .sum()
)

total_outstanding = (
    calling_df["Balance"]
    .sum()
)


# ----------------------------------------------------------
# OPTIONAL VALIDATION VALUES
# ----------------------------------------------------------

duplicate_invoice_count = (
    calling_df["Invoice Number"]
    .duplicated()
    .sum()
)

customer_row_count = len(customer_table)

# ==========================================================
# FASTRANKING COLLECTIONS DASHBOARD
# PART 2 - KPI CARDS & COLLECTIONS QUEUE
# ==========================================================

# ----------------------------------------------------------
# KPI CARD STYLING
# ----------------------------------------------------------

st.markdown(
    """
    <style>
        .kpi-card {
            background-color: rgba(255, 255, 255, 0.04);
            border: 1px solid rgba(255, 255, 255, 0.10);
            border-radius: 12px;
            padding: 18px 16px;
            min-height: 112px;
        }

        .kpi-title {
            font-size: 14px;
            opacity: 0.75;
            margin-bottom: 8px;
        }

        .kpi-value {
            font-size: 25px;
            font-weight: 700;
            line-height: 1.2;
        }
    </style>
    """,
    unsafe_allow_html=True
)


def show_kpi(column, title, value):

    with column:

        st.markdown(
            f"""
            <div class="kpi-card">
                <div class="kpi-title">{title}</div>
                <div class="kpi-value">{value}</div>
            </div>
            """,
            unsafe_allow_html=True
        )


# ==========================================================
# KPI CARDS
# ==========================================================

st.subheader("Collections Overview")

kpi_col_1, kpi_col_2, kpi_col_3 = st.columns(3)

show_kpi(
    kpi_col_1,
    "👥 Customers",
    f"{total_customers:,}"
)

show_kpi(
    kpi_col_2,
    "📄 Outstanding Invoices",
    f"{total_invoices:,}"
)

show_kpi(
    kpi_col_3,
    "💷 Total Outstanding",
    f"£{total_outstanding:,.2f}"
)


kpi_col_4, kpi_col_5, kpi_col_6 = st.columns(3)

show_kpi(
    kpi_col_4,
    "🟢 Current Due",
    f"£{total_current_due:,.2f}"
)

show_kpi(
    kpi_col_5,
    "🔴 Overdue",
    f"£{total_overdue_due:,.2f}"
)

show_kpi(
    kpi_col_6,
    "🟦 Partially Paid Balance",
    f"£{total_partial_due:,.2f}"
)


st.markdown("<br>", unsafe_allow_html=True)


# ==========================================================
# PREPARE DISPLAY TABLE
# ==========================================================

display_table = customer_table.copy()


# ----------------------------------------------------------
# CLEAN CONTACT VALUES
# ----------------------------------------------------------

for col in [
    "Phone_Number",
    "Phone",
    "Mobile",
    "Email"
]:

    display_table[col] = (
        display_table[col]
        .fillna("")
        .astype(str)
        .str.strip()
        .str.replace(
            r"\.0$",
            "",
            regex=True
        )
    )


# ----------------------------------------------------------
# FORMAT DATE COLUMNS
# ----------------------------------------------------------

display_table["Oldest Due Date"] = (
    pd.to_datetime(
        display_table["Oldest_Due_Date"],
        errors="coerce"
    )
    .dt.strftime("%d-%m-%Y")
    .fillna("")
)

display_table["Latest Invoice Date"] = (
    pd.to_datetime(
        display_table["Latest_Invoice_Date"],
        errors="coerce"
    )
    .dt.strftime("%d-%m-%Y")
    .fillna("")
)


# ----------------------------------------------------------
# FORMAT CURRENCY COLUMNS
# ----------------------------------------------------------

display_table["Total Due"] = (
    display_table["Total_Due"]
    .map(lambda value: f"£{value:,.2f}")
)

display_table["Current Due"] = (
    display_table["Current_Due"]
    .map(lambda value: f"£{value:,.2f}")
)

display_table["Overdue Due"] = (
    display_table["Overdue_Due"]
    .map(lambda value: f"£{value:,.2f}")
)


# ----------------------------------------------------------
# RENAME DISPLAY COLUMNS
# ----------------------------------------------------------

display_table = display_table.rename(
    columns={
        "Customer Name": "Customer",
        "Phone_Number": "Primary Number",
        "Invoice_Count": "Invoices",
        "Maximum_Days_Overdue": "Days Overdue"
    }
)


# ----------------------------------------------------------
# SELECT FINAL COLUMN ORDER
# ----------------------------------------------------------

display_table = display_table[
    [
        "Priority",
        "Customer",
        "Primary Number",
        "Phone",
        "Mobile",
        "Email",
        "Invoices",
        "Total Due",
        "Current Due",
        "Overdue Due",
        "Oldest Due Date",
        "Days Overdue",
        "Status",
        "Disposition",
        "PTP Date",
        "Remarks"
    ]
]


# ==========================================================
# DISPLAY COLLECTIONS QUEUE
# ==========================================================

st.subheader("Collections Queue")

st.caption(
    f"{customer_row_count:,} customers | "
    f"{total_invoices:,} unique outstanding invoices"
)

edited_table = st.data_editor(

    display_table,

    width="stretch",

    hide_index=True,

    height=650,

    num_rows="fixed",

    disabled=[
        col for col in display_table.columns
        if col not in [
            "Disposition",
            "PTP Date",
            "Remarks"
        ]
    ],

    column_config={
        "Priority": st.column_config.TextColumn(
            "Priority",
            width="small"
        ),
        "Customer": st.column_config.TextColumn(
            "Customer",
            width="large"
        ),
        "Primary Number": st.column_config.TextColumn(
            "Primary Number",
            width="medium"
        ),
        "Phone": st.column_config.TextColumn(
            "Phone",
            width="medium"
        ),
        "Mobile": st.column_config.TextColumn(
            "Mobile",
            width="medium"
        ),
        "Email": st.column_config.TextColumn(
            "Email",
            width="large"
        ),
        "Invoices": st.column_config.NumberColumn(
            "Invoices",
            format="%d"
        ),
        "Total Due": st.column_config.TextColumn(
            "Total Due",
            width="small"
        ),
        "Current Due": st.column_config.TextColumn(
            "Current Due",
            width="small"
        ),
        "Overdue Due": st.column_config.TextColumn(
            "Overdue Due",
            width="small"
        ),
        "Oldest Due Date": st.column_config.TextColumn(
            "Oldest Due",
            width="small"
        ),
        "Days Overdue": st.column_config.NumberColumn(
            "Days Overdue",
            format="%d"
        ),
        "Status": st.column_config.TextColumn(
            "Status",
            width="small"
        ),
        "Disposition": st.column_config.TextColumn(
            "Disposition",
            width="medium"
        ),
        "PTP Date": st.column_config.TextColumn(
            "PTP Date",
            width="small"
        ),
        "Remarks": st.column_config.TextColumn(
            "Remarks",
            width="large"
        )
    }
)
#####################################################################

## save button

#####################################################################
if st.button(
    "💾 Save Collection Updates",
    type="primary"
):

    save_df = edited_table[
        [
            "Customer",
            "Disposition",
            "PTP Date",
            "Remarks"
        ]
    ].rename(
        columns={
            "Customer":"Customer Name"
        }
    )


    save_collection_updates(
        save_df
    )

    st.success(
        "Collection updates saved successfully."
    )
# ----------------------------------------------------------
# VALIDATION MESSAGE
# ----------------------------------------------------------

if duplicate_invoice_count == 0:

    st.success(
        "Validation passed: no duplicate invoice numbers found."
    )

else:

    st.error(
        f"Validation failed: "
        f"{duplicate_invoice_count:,} duplicate invoice numbers found."
    )
# ==========================================================
# PART 3 - DOWNLOAD COLLECTIONS QUEUE
# ==========================================================

st.divider()

from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ----------------------------------------------------------
# Export dataframe
# ----------------------------------------------------------

export_df = display_table.copy()

output = BytesIO()

with pd.ExcelWriter(
    output,
    engine="openpyxl"
) as writer:

    export_df.to_excel(
        writer,
        sheet_name="Collections Queue",
        index=False
    )

    ws = writer.sheets["Collections Queue"]

    # ------------------------------------------------------
    # Header Formatting
    # ------------------------------------------------------

    header_fill = PatternFill(
        fill_type="solid",
        start_color="1F4E78"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    for cell in ws[1]:

        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    # ------------------------------------------------------
    # Auto Column Width
    # ------------------------------------------------------

    for column_cells in ws.columns:

        length = max(
            len(str(cell.value if cell.value else ""))
            for cell in column_cells
        )

        ws.column_dimensions[
            column_cells[0].column_letter
        ].width = min(length + 3, 40)

# ----------------------------------------------------------
# Download Button
# ----------------------------------------------------------

st.download_button(

    "📥 Download Collections Queue",

    data=output.getvalue(),

    file_name="Collections_Queue.xlsx",

    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

)

# ----------------------------------------------------------
# Footer
# ----------------------------------------------------------

st.caption(
    f"""
    {customer_row_count:,} customers •
    {total_invoices:,} outstanding invoices •
    £{total_outstanding:,.2f} outstanding
    """
)
